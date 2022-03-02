from __future__ import print_function
from PIL import Image,ImageTk
import tkinter as tki
import threading
import cv2,os,sys

from common_utils import *

import pandas as pd
from openpyxl import load_workbook

def rounded(input_list):
    return [round(x,1) for x in input_list]

class camera_gui:
    def __init__(self, cam1,cam2, out_path):
        self.cam1 = cam1
        self.cam2 = cam2
        self.out_path = out_path
        self.frame = None
        self.frame2 = None
        self.root = tki.Tk()
        self.panel = None
        self.panel1 = None
        self.thread = None
        self.stopEvent = None
        self.out_list = []
        self.data_pd = pd.DataFrame()
        self.i , self.i1 = None, None
        btn = tki.Button(self.root, text="Save Data!",command=self.save_results)
        btn.pack(side="bottom", fill="both", expand="yes", padx=10,pady=10)

        self.stopEvent = threading.Event()
        self.thread = threading.Thread(target=self.video_loop, args=())
        self.thread.start()

        self.root.wm_title("ArUCO Detector!")
        self.root.wm_protocol("WM_DELETE_WINDOW", self.onClose)

    
    def video_loop(self):
        with torch.no_grad():
            try:
                while not self.stopEvent.is_set():

                    _,self.frame = self.cam1.read()
                    _,self.frame2 = self.cam2.read()
                    
                    self.frame,self.data = detect_object(self.frame)
                    image = cv2.cvtColor(self.frame,cv2.COLOR_BGR2RGB)               
                    # image,self.data = get_aruco_point(image)
                    self.i = cv2.cvtColor(image.copy(),cv2.COLOR_BGR2RGB)
                    image = Image.fromarray(image)
                    image = ImageTk.PhotoImage(image)

                    self.frame2,self.data1 = detect_object(self.frame2)
                    image2 = cv2.cvtColor(self.frame2,cv2.COLOR_BGR2RGB)
                    # image2,self.data1 = get_aruco_point(image2)
                    # image2 = increase_brightness(image2,50)
                    self.i1 = cv2.cvtColor(image2.copy(),cv2.COLOR_BGR2RGB)
                    image2 = Image.fromarray(image2)
                    image2 = ImageTk.PhotoImage(image2)
                    
                    self.position,self.orientation = get_pose()

                    
                    if self.panel is None:
                        self.panel = tki.Label(image=image)
                        self.panel.image = image
                        self.panel.pack(side="left", padx=10, pady=10)

                        self.panel1 = tki.Label(image=image2)
                        self.panel1.image = image2
                        self.panel1.pack(side="right", padx=10, pady=10)
            
                    # otherwise, simply update the panel
                    else:
                        self.panel.configure(image=image)
                        self.panel.image = image

                        self.panel1.configure(image=image2)
                        self.panel1.image = image2
            except RuntimeError as e:
                print("[INFO] caught a RuntimeError: ",e)

    def save_results(self):
        output = [self.position,self.orientation,self.data,self.data1]
        name = f"/home/krekik/RICAIP/Images/Camera_0_P{rounded(self.position)}_O{rounded(self.orientation)}_A{self.data}.jpg"
        name1 = f"/home/krekik/RICAIP/Images/Camera_1_P{rounded(self.position)}_O{rounded(self.orientation)}_A{self.data1}.jpg"
        if (self.data != []):
            cv2.imwrite(name,self.i)
        if (self.data1 != []):
            cv2.imwrite(name1,self.i1)
        self.out_list.append(output)
        print("Saving results....")
        print(f"| ArUCO: {self.data}| ArUCO: {self.data1} | Position: {self.position} | orientation: {self.orientation} |")
        self.data_pd = pd.DataFrame(self.out_list, columns =['Position', 'Orientation','ArUCO Data 0','ArUCO Data 1'],)

    def onClose(self):
        if os.path.isfile('/home/krekik/RICAIP/Results.xlsx'):
            path = "/home/krekik/RICAIP/Results.xlsx"
            book = load_workbook(path)
            writer = pd.ExcelWriter(path, engine='openpyxl')
            writer.book = book
            writer.sheets = {ws.title: ws for ws in book.worksheets}
            self.data_pd.to_excel(writer, startrow=writer.sheets['Sheet1'].max_row, index = False,header= False)
            writer.save()
        else:
            self.data_pd.to_excel('/home/krekik/RICAIP/Results.xlsx',index = False,header=True)
        print("Exiting gracefully....")
        print("[INFO] closing...")
        self.stopEvent.set()
        print("[INFO] closing...")
        self.cam1.release()
        print("[INFO] closing...")
        self.cam2.release()
        print("[INFO] closing...")
        self.root.destroy()
        print("[INFO] closed...")